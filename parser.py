"""
Parses an ICICI bank statement export and extracts, per transaction, a
best-effort counterparty name/UPI-id and a transaction category, by reading
the bank's free-text transaction narration.

Two ICICI export templates are supported, auto-detected by scanning for a
recognizable header row (not by filename or a fixed row position):

  Format A - "Statement of Account" (SOA) export
      Header row 1, columns: Ac_No, AC_Name, Tran_ID, Tran_Date, Inst_Type,
      Inst_Num, Dr_Amt, Cr_Amt, Balance, Narration, pstd_dt

  Format B - "Detailed Statement" / transaction-history export
      A few title rows, then an account info line
      ("Transactions List - ... -<NAME> (INR) - <ACCOUNT_NO>"), then a
      header row with columns: No., Transaction ID, Value Date,
      Txn Posted Date, ChequeNo., Description, Cr/Dr,
      Transaction Amount(INR), Available Balance(INR)

Both .xlsx (modern) and .xls (legacy) container formats are supported for
either template; the container format is detected from the file's contents.
"""
import re
from datetime import datetime

import openpyxl
import xlrd

FORMAT_A_MARKERS = {"Ac_No", "Tran_Date", "Dr_Amt", "Cr_Amt", "Narration"}
FORMAT_B_MARKERS = {"Transaction ID", "Value Date", "Description", "Cr/Dr", "Transaction Amount(INR)"}

BANK_FEE_PREFIXES = (
    "IMPS Chg", "Chq rtn Chg", "Int on delayed", "REJECT", "EZY",
    "Debit Card", "SMS Charge", "AMC", "GST", "CHEQUE BOOK",
)

# Merchant/card-network narration prefixes that share the same shape:
# PREFIX/MerchantName(padded)/timestamp/reference/city
CARD_MERCHANT_PREFIXES = ("VIN", "VPS", "VSI", "VIR", "IPS")

EXCLUDE_CATEGORIES_FROM_COUNTERPARTY_VIEWS = {
    "Bank Charges / Fees / Returns",
    "Cash Deposit",
}

XLSX_MAGIC = b"PK\x03\x04"       # zip-based: .xlsx / .xlsm
XLS_MAGIC = b"\xd0\xcf\x11\xe0"  # OLE2-based: legacy .xls

FD_LEADING_DIGITS_RE = re.compile(r"^\d{6,}[:\s]")


class StatementFormatError(ValueError):
    """Raised when the uploaded file doesn't look like a supported ICICI export."""


def _clean_name(s):
    if not s:
        return ""
    return s.strip(" -/.")


def _last_nonempty_part(parts):
    for part in reversed(parts):
        cleaned = part.strip()
        if cleaned:
            return cleaned
    return ""


def parse_narration(narration, dr, cr):
    """Return (category, counterparty, direction) for one transaction row."""
    n = (narration or "").strip()
    direction = "IN" if cr and cr > 0 else ("OUT" if dr and dr > 0 else "NONE")

    for pfx in BANK_FEE_PREFIXES:
        if n.startswith(pfx):
            return ("Bank Charges / Fees / Returns", _clean_name(n), direction)

    if n.startswith("ICICI-SUPP"):
        return ("Bank Posting (Interest/Charges)", "ICICI Bank", direction)

    if (n.upper().startswith("TRF TO FD") or n.upper().startswith("TRF FROM FD")
            or "FD CLOS" in n.upper() or FD_LEADING_DIGITS_RE.match(n)):
        return ("Fixed Deposit", "Own Fixed Deposit (FD)", direction)

    parts = n.split("/")

    if n.startswith("UPI/"):
        vpa = parts[3] if len(parts) > 3 else ""
        remark = parts[2] if len(parts) > 2 else ""
        vpa_id = vpa.split("@")[0] if "@" in vpa else vpa
        label = remark if remark and not remark.lower().startswith("payment from ph") else vpa_id
        cp = _clean_name(label) or _clean_name(vpa_id) or "Unknown UPI counterparty"
        return ("UPI", cp, direction)

    if n.startswith("MMT/IMPS/") or n.startswith("MMT/"):
        name = parts[4] if len(parts) > 4 else (parts[3] if len(parts) > 3 else "")
        return ("IMPS", _clean_name(name) or "Unknown IMPS counterparty", direction)

    if n.startswith("INF/"):
        if "self-" in n.lower():
            return ("Internal Fund Transfer (INFT)", "Self / Own Account", direction)
        sub = parts[1].strip() if len(parts) > 1 else ""
        category_map = {"INFT": "Internal Fund Transfer (INFT)", "NEFT": "NEFT", "RTGS": "RTGS"}
        category = category_map.get(sub, "Internal Fund Transfer (INFT)")
        name = _last_nonempty_part(parts[2:]) or _last_nonempty_part(parts)
        return (category, _clean_name(name) or "Unknown", direction)

    if n.startswith("CLG/"):
        name = parts[1] if len(parts) > 1 else ""
        return ("Cheque Clearing (CLG)", _clean_name(name) or "Unknown", direction)

    if n.startswith("TRF/"):
        name = parts[1] if len(parts) > 1 else ""
        return ("Fund Transfer (TRF)", _clean_name(name) or "Unknown", direction)

    if n.startswith("TRFR TO"):
        name = n.split(":", 1)[1] if ":" in n else ""
        return ("Internal Transfer OUT", _clean_name(name) or "Unknown", direction)

    if n.startswith("TRFR FROM"):
        name = n.split(":", 1)[1] if ":" in n else ""
        return ("Internal Transfer IN", _clean_name(name) or "Unknown", direction)

    if n.startswith("NEFT"):
        p = n.replace("NEFT-", "").replace("NEFT/", "").split("-")
        name = p[1] if len(p) > 1 else ""
        return ("NEFT", _clean_name(name) or "Unknown", direction)

    if n.startswith("RTGS"):
        if "-" in n:
            p = n.split("-")
            name = p[2] if len(p) > 2 else ""
        else:
            p = n.split("/")
            name = p[3] if len(p) > 3 else ""
        return ("RTGS", _clean_name(name) or "Unknown", direction)

    if n.startswith("BIL/"):
        sub = parts[1].strip() if len(parts) > 1 else ""
        if sub == "INFT":
            name = _last_nonempty_part(parts)
        else:
            name = parts[3] if len(parts) > 3 else _last_nonempty_part(parts)
        return ("Bill Payment", _clean_name(name) or "Unknown biller", direction)

    if n.startswith("CASH PAID"):
        name = n.split(":", 1)[1] if ":" in n else ""
        return ("Cash Withdrawal/Payment", _clean_name(name) or "Cash", direction)

    if n.startswith("ATM/"):
        return ("ATM Cash Withdrawal" if direction == "OUT" else "ATM Cash Deposit",
                "Self (ATM)", direction)

    if n.startswith("CAM/"):
        if direction == "OUT":
            return ("Cash Withdrawal/Payment", "Self (cash withdrawal)", direction)
        return ("Cash Deposit", "Cash Deposit (depositor not named)", direction)

    if any(parts[0].endswith(p) for p in CARD_MERCHANT_PREFIXES):
        name = parts[1].strip() if len(parts) > 1 else ""
        return ("Card/Merchant Payment", _clean_name(name) or "Unknown merchant", direction)

    if n.startswith("GIB/"):
        name = parts[2].strip() if len(parts) > 2 else ""
        return ("Bill Payment", _clean_name(name) or "Unknown biller", direction)

    if n.startswith("CMS/"):
        name = parts[2].strip() if len(parts) > 2 else _last_nonempty_part(parts)
        return ("Cash Management Service (CMS)", _clean_name(name) or "Unknown", direction)

    if n.startswith("GRS/"):
        return ("Payment Gateway Settlement", "Payment Gateway", direction)

    for ref_prefix in ("VISA REF", "VPS REF"):
        if n.upper().startswith(ref_prefix):
            name = n[len(ref_prefix):].strip()
            return ("Card Payment", _clean_name(name) or "Unknown merchant", direction)

    if n.startswith("ACH/"):
        return ("ACH Transaction", _clean_name(parts[1]) if len(parts) > 1 else "Unknown", direction)

    if n.startswith("IIN/"):
        candidates = [
            x.strip() for x in parts[1:]
            if x.strip() and x.strip().lower() != "i-debit" and not x.strip().isdigit()
        ]
        name = candidates[0] if candidates else ""
        return ("Card/Merchant Payment", _clean_name(name) or "Unknown merchant", direction)

    return ("Other/Unclassified", _clean_name(n) or "Unknown", direction)


def _detect_excel_format(file_stream):
    """Sniff the first bytes to tell modern (.xlsx) from legacy (.xls) Excel files."""
    header = file_stream.read(8)
    file_stream.seek(0)
    if header.startswith(XLSX_MAGIC):
        return "xlsx"
    if header.startswith(XLS_MAGIC):
        return "xls"
    return None


def _all_rows_from_xlsx(file_stream):
    wb = openpyxl.load_workbook(file_stream, data_only=True, read_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _all_rows_from_xls(file_stream):
    wb = xlrd.open_workbook(file_contents=file_stream.read())
    ws = wb.sheet_by_index(0)
    datemode = wb.datemode

    def cell_value(r, c):
        cell = ws.cell(r, c)
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate_as_datetime(cell.value, datemode)
            except (xlrd.XLDateError, ValueError):
                return cell.value
        if cell.ctype == xlrd.XL_CELL_EMPTY:
            return None
        return cell.value

    return [[cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]


def _locate_header(rows, max_scan=25):
    """Find the header row and which known template it matches, scanning the top of the sheet."""
    for i, row in enumerate(rows[:max_scan]):
        cells = {str(c).strip() for c in row if c is not None}
        if FORMAT_A_MARKERS.issubset(cells):
            return i, "A"
        if FORMAT_B_MARKERS.issubset(cells):
            return i, "B"
    return None, None


def _extract_format_b_account_info(rows, header_idx):
    """Scan the title rows above the header for the '...(INR) - <account no>' line."""
    for row in rows[:header_idx]:
        for val in row:
            if not isinstance(val, str) or "(INR)" not in val:
                continue
            left, right = val.split("(INR)", 1)
            name_tokens = [t.strip() for t in left.split("-") if t.strip()]
            name = name_tokens[-1] if name_tokens else None
            m = re.search(r"(\d{6,})", right)
            account_no = m.group(1) if m else None
            if account_no or name:
                return {"account_no": account_no, "account_name": name}
    return {"account_no": None, "account_name": None}


def _parse_date_value(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def _normalize_format_a(rows, header_idx):
    header = [str(h).strip() if h else "" for h in rows[header_idx]]
    col_index = {name: idx for idx, name in enumerate(header)}
    missing = [c for c in ("Ac_No", "AC_Name", "Tran_Date", "Dr_Amt", "Cr_Amt", "Balance", "Narration") if c not in col_index]
    if missing:
        raise StatementFormatError(
            "This doesn't look like a supported ICICI statement export. "
            f"Missing expected column(s): {', '.join(missing)}."
        )

    account_info = {"account_no": None, "account_name": None}
    normalized = []
    for row in rows[header_idx + 1:]:
        if row is None or all(v is None for v in row):
            continue
        if account_info["account_no"] is None:
            ac_no = row[col_index["Ac_No"]]
            ac_name = row[col_index["AC_Name"]]
            if ac_no is not None:
                if isinstance(ac_no, float) and ac_no.is_integer():
                    account_info["account_no"] = str(int(ac_no))
                else:
                    account_info["account_no"] = str(ac_no)
            if ac_name is not None:
                account_info["account_name"] = str(ac_name)
        normalized.append((
            row[col_index["Tran_Date"]],
            row[col_index["Dr_Amt"]] or 0,
            row[col_index["Cr_Amt"]] or 0,
            row[col_index["Balance"]],
            row[col_index["Narration"]],
        ))
    return account_info, normalized


def _normalize_format_b(rows, header_idx):
    header = [str(h).strip() if h else "" for h in rows[header_idx]]
    col_index = {name: idx for idx, name in enumerate(header)}
    missing = [c for c in FORMAT_B_MARKERS if c not in col_index]
    if missing:
        raise StatementFormatError(
            "This doesn't look like a supported ICICI statement export. "
            f"Missing expected column(s): {', '.join(missing)}."
        )

    account_info = _extract_format_b_account_info(rows, header_idx)

    normalized = []
    for row in rows[header_idx + 1:]:
        if row is None or all(v is None for v in row):
            continue
        crdr = row[col_index["Cr/Dr"]]
        amount = row[col_index["Transaction Amount(INR)"]] or 0
        dr = amount if isinstance(crdr, str) and crdr.strip().upper() == "DR" else 0
        cr = amount if isinstance(crdr, str) and crdr.strip().upper() == "CR" else 0
        normalized.append((
            row[col_index["Value Date"]],
            dr,
            cr,
            row[col_index["Available Balance(INR)"]],
            row[col_index["Description"]],
        ))
    return account_info, normalized


def parse_workbook(file_stream):
    """
    Parse an uploaded ICICI statement (a file-like object; .xlsx or legacy .xls,
    and either the SOA or Detailed Statement template, all auto-detected) into
    a list of transaction dicts, plus a small dict of account info (account_no,
    account_name). Raises StatementFormatError if the file isn't recognized.
    """
    excel_fmt = _detect_excel_format(file_stream)
    if excel_fmt == "xlsx":
        rows = _all_rows_from_xlsx(file_stream)
    elif excel_fmt == "xls":
        rows = _all_rows_from_xls(file_stream)
    else:
        raise StatementFormatError(
            "This file doesn't look like a valid Excel file (.xlsx or .xls). "
            "Please upload the original statement export from ICICI, unmodified."
        )

    header_idx, template = _locate_header(rows)
    if template is None:
        raise StatementFormatError(
            "This doesn't look like a supported ICICI statement export "
            "(no recognizable column header was found near the top of the file)."
        )

    if template == "A":
        account_info, normalized = _normalize_format_a(rows, header_idx)
    else:
        account_info, normalized = _normalize_format_b(rows, header_idx)

    results = []
    for tran_date, dr, cr, balance, narration in normalized:
        category, counterparty, direction = parse_narration(narration, dr, cr)
        date_iso = _parse_date_value(tran_date)
        results.append({
            "date": date_iso,
            "category": category,
            "counterparty": counterparty,
            "direction": direction,
            "dr": float(dr) if dr else 0.0,
            "cr": float(cr) if cr else 0.0,
            "balance": float(balance) if balance is not None else None,
            "narration": narration,
        })

    if not results:
        raise StatementFormatError("No transaction rows were found in this file.")

    return results, account_info


def build_summary(transactions):
    """Compute overall totals and per-counterparty aggregates for both directions."""
    def agg_for(direction):
        agg = {}
        for t in transactions:
            if t["direction"] != direction:
                continue
            if t["category"] in EXCLUDE_CATEGORIES_FROM_COUNTERPARTY_VIEWS:
                continue
            key = t["counterparty"]
            entry = agg.setdefault(key, {"name": key, "total": 0.0, "count": 0, "first": t["date"], "last": t["date"]})
            entry["total"] += t["dr"] if direction == "OUT" else t["cr"]
            entry["count"] += 1
            if t["date"] and (entry["first"] is None or t["date"] < entry["first"]):
                entry["first"] = t["date"]
            if t["date"] and (entry["last"] is None or t["date"] > entry["last"]):
                entry["last"] = t["date"]
        return sorted(agg.values(), key=lambda x: -x["total"])

    total_in = sum(t["cr"] for t in transactions if t["direction"] == "IN")
    total_out = sum(t["dr"] for t in transactions if t["direction"] == "OUT")
    dates = sorted(t["date"] for t in transactions if t["date"])

    return {
        "total_transactions": len(transactions),
        "period_start": dates[0] if dates else None,
        "period_end": dates[-1] if dates else None,
        "total_in": total_in,
        "total_out": total_out,
        "net": total_in - total_out,
        "payers": agg_for("IN"),
        "payees": agg_for("OUT"),
    }
